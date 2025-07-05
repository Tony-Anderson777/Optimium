#!/usr/bin/env python3
# coding: utf-8
"""
Optimisation des réservations de salles – version améliorée avec détection de doublons
• Tri intelligent : créneaux les plus courts en premier, puis effectif décroissant
• Algorithmes : Glouton intelligent et Génétique
• Gestion des conflits horaires avec buffer configurable
• Interface multilingue (FR/EN)
• Détection et gestion des doublons par CodeAnalytique
• Correction du décompte des inscrits par CodeAnalytique
"""

import streamlit as st
import pandas as pd
import os
from io import BytesIO
import logging
from typing import List, Tuple, Optional, Dict, Any
from bisect import bisect_left, insort
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import random
import numpy as np
from datetime import datetime, timedelta

st.set_page_config(
    page_title="Optimium l'appli d'optimisation  des réservations de salles du CESI",
    page_icon="🏠",
    layout="wide"
)

# ── LOGGING ──────────────────────────────────────────────────────────
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ── CONSTANTES COLONNES ─────────────────────────────────────────────
NOM_FICHIER_SALLES = "ExtractSalleRouenCESI.xlsx"

# Colonnes salles
COL_NOM_SALLE = "Nom Salle"
COL_CAPACITE = "CapaciteSalle"

# Colonnes réservations
COL_NB_INSCRITS = "NombreInscrit"
COL_SALLE_OLD = "NomSalle"  # Colonne source pour l'ancienne salle
COL_NOM_ANCIENNE_SALLE = "NomAncienneSalle"  # Colonne à créer
COL_DATE = "Date"
COL_HEURE_DEBUT = "Début"
COL_HEURE_FIN = "Fin"
COL_CODE_ANALYTIQUE = "CodeAnalytique"  # Nouvelle colonne pour identifier les promotions

# Colonnes résultats
COL_SALLE_OPTIM = "NomSalle"
COL_TAUX_OCCUP = "TauxOccupation"
COL_CAPACITE_OLD = "CapaciteAncienneSalle"
COL_RAISON_NA = "RaisonNonAttrib"
COL_DEB = "_start"
COL_FIN = "_end"
COL_DUPLICATA = "EstDuplicata"  # Nouvelle colonne pour marquer les doublons

# ── PARAMÈTRES PAR DÉFAUT ───────────────────────────────────────────
SEUIL_BON_DEFAULT = 0.85   # 85%
SEUIL_BAS_DEFAULT = 0.3    # 30%
BUFFER_DEFAULT = 0         # 0 minutes (au lieu de 15)

# ── TRADUCTIONS ─────────────────────────────────────────────────────
LANGS = {
    "fr": {
        "titre": "🎓 Optimium l'appli d'optimisation  des réservations de salles du CESI",
        "description": "Algorithme intelligent d'affectation optimale des salles avec gestion des doublons",
        "salles_chargees": "✅ {n} salles chargées",
        "seuil_bon": "Seuil optimal d'occupation (%)",
        "seuil_bas": "Seuil minimal d'occupation (%)",
        "buffer": "Buffer anti-conflit (minutes)",
        "critere_tri": "Critère de tri principal",
        "critere_duree": "Durée croissante",
        "critere_effectif": "Effectif décroissant",
        "algo": "Algorithme d'optimisation",
        "algo_glouton": "Glouton intelligent",
        "algo_genetique": "Génétique (avancé)",
        "upload": "📤 Fichier de réservations (.xlsx)",
        "attente_fichier": "⏳ En attente du fichier de réservations...",
        "colonnes_manquantes": "❌ Colonnes manquantes : {cols}",
        "apercu_resa": "📋 Aperçu des réservations",
        "optimiser": "🚀 Lancer l'optimisation",
        "resultats": "📊 Résultats de l'optimisation",
        "taux_assign": "Taux d'assignation",
        "taux_moyen": "Taux moyen d'occupation",
        "telecharger": "📥 Télécharger le résultat",
        "non_attrib": "⚠️ {n} réservations non attribuées",
        "reset": "🔄 Réinitialiser",
        "langue": "🌐 Langue",
        "params": "⚙️ Paramètres",
        "stats": "📈 Statistiques",
        "erreur_salles": "❌ Impossible de charger le fichier des salles",
        "erreur_optim": "❌ Erreur lors de l'optimisation",
        "generations": "Générations (algo génétique)",
        "population": "Taille population",
        "mutation": "Taux de mutation (%)",
        "doublons": "🔎 Doublons détectés",
        "inscrits_corriges": "👥 Inscrits corrigés (sans doublons)"
    },
    "en": {
        "titre": "🎓 Optimium, CESI's room reservation optimization application",
        "description": "Smart algorithm for optimal room assignment with duplicate handling",
        "salles_chargees": "✅ {n} rooms loaded",
        "seuil_bon": "Optimal occupation threshold (%)",
        "seuil_bas": "Minimal occupation threshold (%)",
        "buffer": "Anti-conflict buffer (minutes)",
        "critere_tri": "Main sorting criterion",
        "critere_duree": "Duration ascending",
        "critere_effectif": "Headcount descending",
        "algo": "Optimization algorithm",
        "algo_glouton": "Smart Greedy",
        "algo_genetique": "Genetic (advanced)",
        "upload": "📤 Booking file (.xlsx)",
        "attente_fichier": "⏳ Waiting for booking file...",
        "colonnes_manquantes": "❌ Missing columns: {cols}",
        "apercu_resa": "📋 Bookings preview",
        "optimiser": "🚀 Start optimization",
        "resultats": "📊 Optimization results",
        "taux_assign": "Assignment rate",
        "taux_moyen": "Average occupation rate",
        "telecharger": "📥 Download results",
        "non_attrib": "⚠️ {n} bookings not assigned",
        "reset": "🔄 Reset",
        "langue": "🌐 Language",
        "params": "⚙️ Parameters",
        "stats": "📈 Statistics",
        "erreur_salles": "❌ Cannot load rooms file",
        "erreur_optim": "❌ Optimization error",
        "generations": "Generations (genetic algo)",
        "population": "Population size",
        "mutation": "Mutation rate (%)",
        "doublons": "🔎 Duplicates detected",
        "inscrits_corriges": "👥 Corrected headcount (no duplicates)"
    }
}

# ── FONCTIONS UTILITAIRES ───────────────────────────────────────────
def normaliser_nom_salle(nom: str) -> str:
    """Normalise le nom d'une salle pour la correspondance."""
    if pd.isna(nom) or nom == "":
        return ""
    nom_clean = str(nom).strip().upper()
    # Supprimer les caractères spéciaux et espaces multiples
    import re
    nom_clean = re.sub(r'[^\w\s]', '', nom_clean)
    nom_clean = re.sub(r'\s+', ' ', nom_clean).strip()
    return nom_clean

def salle_libre(planning: List[Tuple], start: datetime, end: datetime, buffer_min: int = 0) -> bool:
    """Vérifie si une salle est libre pour un créneau donné avec buffer."""
    if not planning:
        return True
    
    buffer = timedelta(minutes=buffer_min)
    start_buf = start - buffer
    end_buf = end + buffer
    
    idx = bisect_left(planning, (start_buf, start_buf))
    
    for i in range(max(0, idx-1), min(len(planning), idx+2)):
        existing_start, existing_end = planning[i]
        if not (end_buf <= existing_start or start_buf >= existing_end):
            return False
    return True

def calculer_score_fitness(individu: List[str], df_resa: pd.DataFrame, 
                          cap_lookup: Dict[str, int], seuil_bon: float, 
                          seuil_bas: float) -> float:
    """Calcule le score de fitness pour l'algorithme génétique."""
    score = 0
    penalite = 0
    
    for idx, salle in enumerate(individu):
        if salle is None or salle == "Aucune salle adaptée":
            penalite += 100
            continue
            
        inscrits = df_resa.iloc[idx][COL_NB_INSCRITS]
        if pd.isna(inscrits) or inscrits <= 0:
            penalite += 50
            continue
            
        cap = cap_lookup.get(salle, 0)
        if cap == 0 or inscrits > cap:
            penalite += 75
            continue
            
        taux = inscrits / cap
        
        if seuil_bas <= taux <= seuil_bon:
            score += taux * 100
        elif taux > seuil_bon:
            score += seuil_bon * 100 - (taux - seuil_bon) * 50
        else:
            score += taux * 50
    
    return score - penalite

# ── DÉTECTION DES DOUBLONS ──────────────────────────────────────────
def detecter_doublons(df_resa: pd.DataFrame) -> pd.DataFrame:
    """Identifie les réservations exactement identiques (même CodeAnalytique, Date, Heure, Inscrits)."""
    df = df_resa.copy()
    
    df[COL_DUPLICATA] = False
    
    # Regrouper par CodeAnalytique, Date, Heure de début, Heure de fin et Nombre d'inscrits
    grouped = df.groupby([COL_CODE_ANALYTIQUE, COL_DATE, COL_HEURE_DEBUT, COL_HEURE_FIN, COL_NB_INSCRITS])
    
    for (code, date, debut, fin, inscrits), group in grouped:
        if len(group) > 1:  # Plus d'une réservation identique
            # Marquer toutes les doublons sauf la première
            for i, (idx, row) in enumerate(group.iterrows()):
                if i > 0:  # Garder la première, marquer les autres comme doublons
                    df.loc[idx, COL_DUPLICATA] = True
    
    return df

# ── CORRECTION DU COMPTE DES INSCRITS ───────────────────────────────
def corriger_inscrits(df_resa: pd.DataFrame) -> int:
    """Calcule le nombre total d'inscrits uniques par CodeAnalytique."""
    if COL_CODE_ANALYTIQUE not in df_resa.columns:
        return df_resa[COL_NB_INSCRITS].sum()
    
    # Prendre le max des inscrits pour chaque CodeAnalytique
    return df_resa.groupby(COL_CODE_ANALYTIQUE)[COL_NB_INSCRITS].max().sum()

# ── CHARGEMENT DES SALLES ───────────────────────────────────────────
@st.cache_data
def charger_salles(path: str) -> Optional[pd.DataFrame]:
    """Charge et nettoie le catalogue des salles."""
    try:
        if not os.path.exists(path):
            st.error(f"❌ Fichier salles introuvable : {path}")
            return None
            
        df = pd.read_excel(path)
        logger.info(f"Salles chargées : {len(df)} lignes")
        
        df[COL_CAPACITE] = pd.to_numeric(df[COL_CAPACITE], errors="coerce")
        df = df.dropna(subset=[COL_NOM_SALLE, COL_CAPACITE])
        df[COL_NOM_SALLE] = df[COL_NOM_SALLE].astype(str).str.strip()
        
        if (df[COL_CAPACITE] <= 0).any():
            st.warning("⚠️ Capacités négatives détectées et supprimées")
            df = df[df[COL_CAPACITE] > 0]
        
        # Gérer les doublons en gardant la capacité la plus élevée
        df = df.sort_values(COL_CAPACITE, ascending=False).drop_duplicates(subset=[COL_NOM_SALLE], keep='first')
        
        logger.info(f"Salles chargées après nettoyage : {len(df)} lignes")
        logger.info(f"Capacités uniques : {df[COL_CAPACITE].unique()}")
        
        return df.sort_values(COL_CAPACITE).reset_index(drop=True)
        
    except Exception as e:
        logger.error(f"Erreur chargement salles : {e}")
        st.error(f"❌ Erreur chargement salles : {e}")
        return None

# ── ALGORITHME GLOUTON INTELLIGENT ──────────────────────────────────
def optimiser_glouton(df_resa: pd.DataFrame, df_salles: pd.DataFrame,
                     seuil_bon: float, seuil_bas: float, 
                     buffer_min: int = 15) -> Optional[pd.DataFrame]:
    """Algorithme glouton intelligent avec gestion des doublons."""
    try:
        cap_lookup = dict(zip(df_salles[COL_NOM_SALLE], df_salles[COL_CAPACITE]))
        # Créer un lookup normalisé pour la correspondance des salles
        cap_lookup_normalise = {normaliser_nom_salle(nom): cap for nom, cap in cap_lookup.items()}
        
        # Debug: Vérifier les doublons dans le catalogue des salles
        salles_dupliquees = df_salles[df_salles[COL_NOM_SALLE].duplicated(keep=False)]
        if not salles_dupliquees.empty:
            st.warning(f"⚠️ Salles dupliquées dans le catalogue: {salles_dupliquees[COL_NOM_SALLE].tolist()}")
            st.dataframe(salles_dupliquees[[COL_NOM_SALLE, COL_CAPACITE]])
        
        # Debug: Afficher toutes les capacités pour S158
        salles_s158 = df_salles[df_salles[COL_NOM_SALLE].str.contains('S158', case=False, na=False)]
        if not salles_s158.empty:
            st.info(f"🔍 Capacités trouvées pour S158: {salles_s158[[COL_NOM_SALLE, COL_CAPACITE]].to_dict('records')}")
        
        df = detecter_doublons(df_resa)
        
        if COL_CAPACITE in df.columns:
            df = df.drop(columns=[COL_CAPACITE])
        
        df[COL_DEB] = pd.to_datetime(
            df[COL_DATE].astype(str) + " " + df[COL_HEURE_DEBUT].astype(str),
            dayfirst=True, errors="coerce"
        )
        df[COL_FIN] = pd.to_datetime(
            df[COL_DATE].astype(str) + " " + df[COL_HEURE_FIN].astype(str),
            dayfirst=True, errors="coerce"
        )
        df[COL_NB_INSCRITS] = pd.to_numeric(df[COL_NB_INSCRITS], errors="coerce")
        
        df[COL_DUPLICATA] = False
        
        mask_valid = (
            df[COL_DEB].notna() & 
            df[COL_FIN].notna() & 
            (df[COL_FIN] > df[COL_DEB]) &
            df[COL_NB_INSCRITS].notna() & 
            (df[COL_NB_INSCRITS] > 0) &
            (df[COL_DUPLICATA] == False)  # Exclure les doublons
        )
        
        df_valid = (
            df[mask_valid]
            .sort_values([COL_DEB, COL_NB_INSCRITS],
                        ascending=[True, False])
            .copy()
        )
        df_invalid = df[~mask_valid].copy()
        
        plannings = {salle: [] for salle in df_salles[COL_NOM_SALLE]}
        results = []
        
        for _, row in df_valid.iterrows():
            inscrits = row[COL_NB_INSCRITS]
            start, end = row[COL_DEB], row[COL_FIN]
            
            # Récupérer la salle actuellement assignée (ancienne salle)
            old_room = str(row.get(COL_NOM_ANCIENNE_SALLE, "")).strip()
            
            # Normaliser le nom de la salle pour la correspondance
            old_room_normalise = normaliser_nom_salle(old_room)
            
            # Debug: Afficher les informations de débogage seulement si problème
            if old_room_normalise and old_room_normalise not in cap_lookup_normalise:
                st.warning(f"⚠️ Salle '{old_room}' (normalisée: '{old_room_normalise}') non trouvée dans le catalogue")
                st.info(f"Salles disponibles (normalisées): {list(cap_lookup_normalise.keys())[:5]}...")
            elif old_room_normalise:
                capacite_trouvee = cap_lookup_normalise.get(old_room_normalise, 'N/A')
                st.success(f"✅ Salle '{old_room}' (normalisée: '{old_room_normalise}') trouvée avec capacité: {capacite_trouvee}")
                
                # Debug: Afficher toutes les entrées pour cette salle dans le catalogue
                salles_correspondantes = df_salles[df_salles[COL_NOM_SALLE].str.contains(old_room, case=False, na=False)]
                if len(salles_correspondantes) > 1:
                    st.warning(f"⚠️ Plusieurs entrées trouvées pour '{old_room}': {salles_correspondantes[[COL_NOM_SALLE, COL_CAPACITE]].to_dict('records')}")
            
            cap_old = cap_lookup_normalise.get(old_room_normalise, pd.NA)
            
            best_room = None
            best_ratio = -1.0
            found_adequate = False
            found_busy = False
            
            for _, salle_row in df_salles.sort_values(COL_CAPACITE).iterrows():
                salle_nom = salle_row[COL_NOM_SALLE]
                cap = salle_row[COL_CAPACITE]
                
                if cap >= inscrits:
                    found_adequate = True
                    if salle_libre(plannings[salle_nom], start, end, buffer_min):
                        ratio = inscrits / cap
                        if ratio > best_ratio:
                            best_room, best_ratio = salle_nom, ratio
                    else:
                        found_busy = True
            
            raison = pd.NA
            if best_room is None:
                if not found_adequate:
                    raison = "Capacité insuffisante"
                elif found_busy:
                    raison = "Conflit horaire"
                else:
                    raison = "Erreur allocation"
                best_room = "Aucune salle adaptée"
                capacite_assignee = pd.NA
                best_ratio = pd.NA
            else:
                insort(plannings[best_room], (start, end))
                capacite_assignee = cap_lookup[best_room]
                
                if best_ratio >= seuil_bon:
                    raison = f"Taux optimal ({best_ratio:.0%})"
                elif best_ratio <= seuil_bas:
                    raison = f"Sous-utilisé ({best_ratio:.0%})"
            
            result_row = row.to_dict()
            result_row.update({
                COL_SALLE_OPTIM: best_room,
                COL_CAPACITE: capacite_assignee,
                COL_TAUX_OCCUP: best_ratio,
                COL_CAPACITE_OLD: cap_old,
                COL_RAISON_NA: raison
            })
            
            # S'assurer que la colonne NomAncienneSalle est présente
            if COL_NOM_ANCIENNE_SALLE not in result_row:
                result_row[COL_NOM_ANCIENNE_SALLE] = old_room
            results.append(result_row)
        
        df_final = pd.concat([
            pd.DataFrame(results), 
            df_invalid
        ], ignore_index=True)
        
        df_final.drop(columns=[COL_DEB, COL_FIN], 
                     inplace=True, errors="ignore")
        
        def reorganiser_colonnes(cols):
            # Placer CapaciteAncienneSalle à côté de NomAncienneSalle
            if COL_NOM_ANCIENNE_SALLE in cols and COL_CAPACITE_OLD in cols:
                idx = cols.index(COL_NOM_ANCIENNE_SALLE) + 1
                cols.insert(idx, cols.pop(cols.index(COL_CAPACITE_OLD)))
            
            # Placer CapaciteSalle à côté de NomSalle
            if COL_SALLE_OPTIM in cols and COL_CAPACITE in cols:
                idx = cols.index(COL_SALLE_OPTIM) + 1
                cols.insert(idx, cols.pop(cols.index(COL_CAPACITE)))
            
            return cols
        
        df_final = df_final[reorganiser_colonnes(list(df_final.columns))]
        return df_final
        
    except Exception as e:
        logger.error(f"Erreur optimisation glouton : {e}")
        st.error(f"❌ Erreur optimisation : {e}")
        return None

# ── ALGORITHME GÉNÉTIQUE ────────────────────────────────────────────
def optimiser_genetique(df_resa: pd.DataFrame, df_salles: pd.DataFrame,
                       seuil_bon: float, seuil_bas: float,
                       generations: int = 100, population_size: int = 50,
                       mutation_rate: float = 0.1) -> Optional[pd.DataFrame]:
    """Algorithme génétique pour l'optimisation globale avec gestion des doublons."""
    try:
        salles = list(df_salles[COL_NOM_SALLE])
        cap_lookup = dict(zip(df_salles[COL_NOM_SALLE], df_salles[COL_CAPACITE]))
        # Créer un lookup normalisé pour la correspondance des salles
        cap_lookup_normalise = {normaliser_nom_salle(nom): cap for nom, cap in cap_lookup.items()}
        
        df = detecter_doublons(df_resa)
        df = df[df[COL_DUPLICATA] == False].copy()  # Exclure les doublons
        
        df[COL_NB_INSCRITS] = pd.to_numeric(df[COL_NB_INSCRITS], errors="coerce")
        
        n_resa = len(df)
        
        def generer_individu():
            individu = []
            for _, row in df.iterrows():
                inscrits = row[COL_NB_INSCRITS]
                if pd.isna(inscrits) or inscrits <= 0:
                    individu.append(None)
                else:
                    salles_adequates = [s for s in salles if cap_lookup[s] >= inscrits]
                    if salles_adequates:
                        individu.append(random.choice(salles_adequates))
                    else:
                        individu.append(None)
            return individu
        
        def croiser(parent1, parent2):
            point_coupure = random.randint(1, n_resa - 1)
            return parent1[:point_coupure] + parent2[point_coupure:]
        
        def muter(individu):
            individu_mute = individu.copy()
            for i in range(n_resa):
                if random.random() < mutation_rate:
                    inscrits = df.iloc[i][COL_NB_INSCRITS]
                    if not pd.isna(inscrits) and inscrits > 0:
                        salles_adequates = [s for s in salles if cap_lookup[s] >= inscrits]
                        if salles_adequates:
                            individu_mute[i] = random.choice(salles_adequates)
            return individu_mute
        
        population = [generer_individu() for _ in range(population_size)]
        
        progress_bar = st.progress(0)
        for gen in range(generations):
            scores = [calculer_score_fitness(ind, df, cap_lookup, seuil_bon, seuil_bas) 
                     for ind in population]
            
            couples_score_pop = list(zip(scores, population))
            couples_score_pop.sort(key=lambda x: x[0], reverse=True)
            
            elite_size = population_size // 5
            nouvelle_pop = [ind for _, ind in couples_score_pop[:elite_size]]
            
            while len(nouvelle_pop) < population_size:
                parents = random.sample(couples_score_pop[:population_size//2], 2)
                parent1, parent2 = parents[0][1], parents[1][1]
                
                enfant = croiser(parent1, parent2)
                enfant = muter(enfant)
                nouvelle_pop.append(enfant)
            
            population = nouvelle_pop
            progress_bar.progress((gen + 1) / generations)
        
        scores_finaux = [calculer_score_fitness(ind, df, cap_lookup, seuil_bon, seuil_bas) 
                        for ind in population]
        meilleur_individu = population[scores_finaux.index(max(scores_finaux))]
        
        results = []
        for i, salle in enumerate(meilleur_individu):
            row = df.iloc[i].to_dict()
            inscrits = row[COL_NB_INSCRITS]
            
            if salle is None or pd.isna(inscrits) or inscrits <= 0:
                salle_finale = "Aucune salle adaptée"
                cap = pd.NA
                taux = pd.NA
                raison = "Non attribuée"
            else:
                cap = cap_lookup.get(salle, 0)
                if cap == 0 or inscrits > cap:
                    salle_finale = "Aucune salle adaptée"
                    cap = pd.NA
                    taux = pd.NA
                    raison = "Capacité insuffisante"
                else:
                    salle_finale = salle
                    taux = inscrits / cap
                    if taux >= seuil_bon:
                        raison = f"Taux optimal ({taux:.0%})"
                    elif taux <= seuil_bas:
                        raison = f"Sous-utilisé ({taux:.0%})"
                    else:
                        raison = pd.NA
            
            # Récupérer la salle actuellement assignée (ancienne salle) pour l'algorithme génétique
            old_room_genetic = str(row.get(COL_NOM_ANCIENNE_SALLE, "")).strip()
            old_room_genetic_normalise = normaliser_nom_salle(old_room_genetic)
            if old_room_genetic_normalise and old_room_genetic_normalise not in cap_lookup_normalise:
                st.warning(f"⚠️ Algo génétique: Salle '{old_room_genetic}' (normalisée: '{old_room_genetic_normalise}') non trouvée")
            
            row.update({
                COL_SALLE_OPTIM: salle_finale,
                COL_CAPACITE: cap,
                COL_TAUX_OCCUP: taux,
                COL_CAPACITE_OLD: cap_lookup_normalise.get(old_room_genetic_normalise, pd.NA),
                COL_RAISON_NA: raison
            })
            
            # S'assurer que la colonne NomAncienneSalle est présente
            if COL_NOM_ANCIENNE_SALLE not in row:
                row[COL_NOM_ANCIENNE_SALLE] = old_room_genetic
            results.append(row)
        
        progress_bar.empty()
        df_final = pd.DataFrame(results)
        df_invalid = df_resa[df_resa[COL_DUPLICATA] == True].copy()
        df_final = pd.concat([df_final, df_invalid], ignore_index=True)
        return df_final
        
    except Exception as e:
        logger.error(f"Erreur algorithme génétique : {e}")
        st.error(f"❌ Erreur algorithme génétique : {e}")
        return None

# ── EXPORT EXCEL AVEC FORMATAGE ─────────────────────────────────────
def exporter_excel(df: pd.DataFrame, seuil_bon: float, seuil_bas: float) -> BytesIO:
    """Exporte vers Excel avec formatage conditionnel."""
    buffer = BytesIO()
    
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Optimisation")
        
        ws = writer.sheets["Optimisation"]
        
        # Appliquer le format JJ/MM/AAAA à la colonne Date si elle existe
        if COL_DATE in df.columns:
            col_date_idx = df.columns.get_loc(COL_DATE) + 1
            col_lettre_date = get_column_letter(col_date_idx)
            for cell in ws[col_lettre_date][1:]:
                if cell.value is not None:
                    cell.number_format = "DD/MM/YYYY"
        
        if COL_TAUX_OCCUP in df.columns:
            col_taux = df.columns.get_loc(COL_TAUX_OCCUP) + 1
            col_lettre = get_column_letter(col_taux)
            
            for cell in ws[col_lettre][1:]:
                if cell.value is not None:
                    cell.number_format = "0.00%"
        
        if COL_DUPLICATA in df.columns:
            col_duplicata = df.columns.get_loc(COL_DUPLICATA) + 1
            col_lettre_duplicata = get_column_letter(col_duplicata)
            plage_duplicata = f"{col_lettre_duplicata}2:{col_lettre_duplicata}{len(df)+1}"
            ws.conditional_formatting.add(
                plage_duplicata,
                CellIsRule(
                    operator="equal",
                    formula=["TRUE"],
                    fill=PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                )
            )
        
        for col_idx, column_cells in enumerate(ws.columns, 1):
            max_length = max(
                len(str(cell.value)) if cell.value is not None else 0 
                for cell in column_cells
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
        
        if COL_TAUX_OCCUP in df.columns:
            plage = f"{col_lettre}2:{col_lettre}{len(df)+1}"
            ws.conditional_formatting.add(
                plage,
                CellIsRule(
                    operator="greaterThanOrEqual",
                    formula=[str(seuil_bon)],
                    fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                )
            )
            ws.conditional_formatting.add(
                plage,
                CellIsRule(
                    operator="lessThanOrEqual", 
                    formula=[str(seuil_bas)],
                    fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                )
            )
        
        # Appliquer le format HH:MM aux colonnes d'heures
        for col_heure in [COL_HEURE_DEBUT, COL_HEURE_FIN]:
            if col_heure in df.columns:
                col_heure_idx = df.columns.get_loc(col_heure) + 1
                col_lettre_heure = get_column_letter(col_heure_idx)
                for cell in ws[col_lettre_heure][1:]:
                    if cell.value is not None:
                        cell.number_format = "HH:MM"
    
    return buffer

# ── INTERFACE STREAMLIT PRINCIPALE ──────────────────────────────────
def main():
    if "langue" not in st.session_state:
        st.session_state.langue = "fr"
    
    with st.sidebar:
        st.header(LANGS[st.session_state.langue]["params"])
        
        nouvelle_langue = st.selectbox(
            LANGS[st.session_state.langue]["langue"],
            options=["fr", "en"],
            index=0 if st.session_state.langue == "fr" else 1
        )
        if nouvelle_langue != st.session_state.langue:
            st.session_state.langue = nouvelle_langue
            st.rerun()
        
        t = LANGS[st.session_state.langue]
        
        if st.button(t["reset"]):
            st.cache_data.clear()
            st.rerun()
    
    st.title(t["titre"])
    st.markdown(f"*{t['description']}*")
    
    path_salles = os.path.join(os.path.dirname(__file__), NOM_FICHIER_SALLES)
    df_salles = charger_salles(path_salles)
    
    if df_salles is None:
        st.error(t["erreur_salles"])
        st.stop()
    
    with st.sidebar:
        st.success(t["salles_chargees"].format(n=len(df_salles)))
        
        # Debug: Afficher les salles du catalogue
        st.info(f"🏠 Salles du catalogue: {list(df_salles[COL_NOM_SALLE])}")
        
        with st.expander("📋 Salles disponibles"):
            st.dataframe(df_salles, hide_index=True, use_container_width=True)
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("⚙️ Paramètres d'optimisation")
        
        seuil_bon = st.slider(
            t["seuil_bon"],
            min_value=50, max_value=100, value=int(SEUIL_BON_DEFAULT * 100), step=5
        ) / 100
        
        seuil_bas = st.slider(
            t["seuil_bas"], 
            min_value=0, max_value=50, value=int(SEUIL_BAS_DEFAULT * 100), step=5
        ) / 100
        
        buffer_minutes = st.slider(
            t["buffer"],
            min_value=0, max_value=60, value=BUFFER_DEFAULT, step=5
        )
    
    with col2:
        st.subheader("🤖 Algorithme")
        
        algo_choisi = st.selectbox(
            t["algo"],
            options=[t["algo_glouton"], t["algo_genetique"]]
        )
        
        if algo_choisi == t["algo_genetique"]:
            generations = st.slider(t["generations"], 50, 300, 100, 25)
            population = st.slider(t["population"], 20, 100, 50, 10) 
            mutation = st.slider(t["mutation"], 5, 30, 10, 5) / 100
    
    st.subheader("📤 Import des données")
    fichier_resa = st.file_uploader(t["upload"], type="xlsx")
    
    if fichier_resa is None:
        st.info(t["attente_fichier"])
        return
    
    try:
        df_resa = pd.read_excel(fichier_resa)
        
    except Exception as e:
        st.error(f"❌ Erreur lecture fichier : {e}")
        return
    
    colonnes_requises = [COL_DATE, COL_HEURE_DEBUT, COL_HEURE_FIN, COL_NB_INSCRITS, COL_CODE_ANALYTIQUE, COL_SALLE_OLD]
    colonnes_manquantes = [col for col in colonnes_requises if col not in df_resa.columns]
    
    if colonnes_manquantes:
        st.error(t["colonnes_manquantes"].format(cols=", ".join(colonnes_manquantes)))
        return
    
    df_resa = detecter_doublons(df_resa)
    
    # Créer la colonne NomAncienneSalle à partir de NomSalle
    if COL_SALLE_OLD in df_resa.columns:
        df_resa[COL_NOM_ANCIENNE_SALLE] = df_resa[COL_SALLE_OLD]
    
    # Debug: Afficher les salles dans le fichier de réservations
    if COL_NOM_ANCIENNE_SALLE in df_resa.columns:
        salles_resa = df_resa[COL_NOM_ANCIENNE_SALLE].unique()
        st.info(f"📋 Salles dans les réservations: {list(salles_resa)}")
        
        # Vérifier les correspondances avec normalisation
        salles_catalogue = set(df_salles[COL_NOM_SALLE])
        salles_catalogue_normalise = {normaliser_nom_salle(s) for s in salles_catalogue}
        salles_resa_normalise = {normaliser_nom_salle(s) for s in salles_resa}
        
        # Créer un rapport de correspondance
        st.subheader("🔍 Rapport de correspondance des salles")
        
        correspondances = []
        for salle_resa in salles_resa:
            salle_normalise = normaliser_nom_salle(salle_resa)
            if salle_normalise in salles_catalogue_normalise:
                # Trouver la salle originale dans le catalogue
                salle_catalogue = None
                for salle_cat in salles_catalogue:
                    if normaliser_nom_salle(salle_cat) == salle_normalise:
                        salle_catalogue = salle_cat
                        break
                capacite = df_salles[df_salles[COL_NOM_SALLE] == salle_catalogue][COL_CAPACITE].iloc[0] if salle_catalogue else "N/A"
                correspondances.append({
                    "Salle (réservations)": salle_resa,
                    "Salle (catalogue)": salle_catalogue,
                    "Capacité": capacite,
                    "Statut": "✅ Correspondance"
                })
            else:
                correspondances.append({
                    "Salle (réservations)": salle_resa,
                    "Salle (catalogue)": "Non trouvée",
                    "Capacité": "N/A",
                    "Statut": "❌ Non trouvée"
                })
        
        df_correspondances = pd.DataFrame(correspondances)
        st.dataframe(df_correspondances, use_container_width=True)
        
        # Afficher un rapport détaillé des capacités
        st.subheader("📊 Rapport détaillé des capacités")
        
        capacites_detail = []
        for salle_resa in salles_resa:
            salle_normalise = normaliser_nom_salle(salle_resa)
            salles_correspondantes = df_salles[df_salles[COL_NOM_SALLE].str.contains(salle_resa, case=False, na=False)]
            
            if not salles_correspondantes.empty:
                for _, row in salles_correspondantes.iterrows():
                    capacites_detail.append({
                        "Salle (réservations)": salle_resa,
                        "Salle (catalogue)": row[COL_NOM_SALLE],
                        "Capacité": row[COL_CAPACITE],
                        "Normalisée": salle_normalise
                    })
            else:
                capacites_detail.append({
                    "Salle (réservations)": salle_resa,
                    "Salle (catalogue)": "Non trouvée",
                    "Capacité": "N/A",
                    "Normalisée": salle_normalise
                })
        
        df_capacites = pd.DataFrame(capacites_detail)
        st.dataframe(df_capacites, use_container_width=True)
    
    st.subheader(t["apercu_resa"])
    st.dataframe(df_resa.head(10), use_container_width=True)
    
    with st.expander("📊 Statistiques des données"):
        col_stat1, col_stat2, col_stat3 = st.columns(3)
        with col_stat1:
            st.metric("📋 Réservations", len(df_resa))
        with col_stat2:
            inscrits_corriges = corriger_inscrits(df_resa)
            st.metric(t["inscrits_corriges"], f"{inscrits_corriges:,}")
        with col_stat3:
            capacite_totale = df_salles[COL_CAPACITE].sum()
            st.metric("🏠 Capacité totale", f"{capacite_totale:,}")
        
        nb_doublons = df_resa[COL_DUPLICATA].sum()
        if nb_doublons > 0:
            st.warning(t["doublons"].format(n=nb_doublons))
    
    st.markdown("---")
    
    if st.button(t["optimiser"], type="primary", use_container_width=True):
        with st.spinner("🔄 Optimisation en cours..."):
            start_time = datetime.now()
            
            if algo_choisi == t["algo_glouton"]:
                df_optimise = optimiser_glouton(
                    df_resa, df_salles, seuil_bon, seuil_bas, buffer_minutes
                )
            else:
                df_optimise = optimiser_genetique(
                    df_resa, df_salles, seuil_bon, seuil_bas,
                    generations, population, mutation
                )
            
            end_time = datetime.now()
            temps_execution = (end_time - start_time).total_seconds()
        
        if df_optimise is not None:
            st.success(f"✅ Optimisation terminée en {temps_execution:.2f}s")
            
            st.subheader(t["resultats"])
            
            reservations_attribuees = df_optimise[
                df_optimise[COL_SALLE_OPTIM] != "Aucune salle adaptée"
            ]
            
            col_met1, col_met2, col_met3, col_met4 = st.columns(4)
            
            with col_met1:
                taux_assignation = len(reservations_attribuees) / len(df_optimise) * 100
                st.metric(
                    t["taux_assign"], 
                    f"{taux_assignation:.1f}%",
                    delta=f"{len(reservations_attribuees)}/{len(df_optimise)}"
                )
            
            with col_met2:
                if not reservations_attribuees.empty:
                    taux_moyen = reservations_attribuees[COL_TAUX_OCCUP].mean()
                    st.metric(t["taux_moyen"], f"{taux_moyen:.1%}")
                else:
                    st.metric(t["taux_moyen"], "N/A")
            
            with col_met3:
                nb_non_attrib = len(df_optimise) - len(reservations_attribuees)
                st.metric("❌ Non attribuées", nb_non_attrib)
            
            with col_met4:
                if not reservations_attribuees.empty:
                    salles_utilisees = reservations_attribuees[COL_SALLE_OPTIM].nunique()
                    st.metric("🏠 Salles utilisées", f"{salles_utilisees}/{len(df_salles)}")
                else:
                    st.metric("🏠 Salles utilisées", "0")
            
            st.dataframe(df_optimise, use_container_width=True, hide_index=True)
            
            if not reservations_attribuees.empty:
                st.subheader("📈 Analyses détaillées")
                
                tab1, tab2, tab3 = st.tabs(["Distribution des taux", "Utilisation des salles", "Diagnostics"])
                
                with tab1:
                    taux_valides = reservations_attribuees[COL_TAUX_OCCUP].dropna()
                    if not taux_valides.empty:
                        st.bar_chart(taux_valides.value_counts().sort_index())
                        
                        col_stats1, col_stats2 = st.columns(2)
                        with col_stats1:
                            st.metric("📊 Taux médian", f"{taux_valides.median():.1%}")
                            st.metric("📈 Taux maximum", f"{taux_valides.max():.1%}")
                        with col_stats2:
                            nb_optimaux = (taux_valides >= seuil_bon).sum()
                            st.metric("✅ Taux optimaux", f"{nb_optimaux} ({nb_optimaux/len(taux_valides)*100:.1f}%)")
                            nb_sous_utilises = (taux_valides <= seuil_bas).sum()
                            st.metric("⚠️ Sous-utilisés", f"{nb_sous_utilises} ({nb_sous_utilises/len(taux_valides)*100:.1f}%)")
                
                with tab2:
                    utilisation_salles = reservations_attribuees[COL_SALLE_OPTIM].value_counts()
                    st.bar_chart(utilisation_salles)
                    
                    st.subheader("🏆 Top 5 des salles les plus utilisées")
                    for i, (salle, count) in enumerate(utilisation_salles.head().items(), 1):
                        capacite = df_salles[df_salles[COL_NOM_SALLE] == salle][COL_CAPACITE].iloc[0]
                        st.write(f"{i}. **{salle}** - {count} réservations (capacité: {capacite})")
                
                with tab3:
                    raisons = df_optimise[COL_RAISON_NA].value_counts()
                    if not raisons.empty:
                        st.subheader("🔍 Diagnostics")
                        for raison, count in raisons.items():
                            if pd.notna(raison):
                                st.write(f"• **{raison}** : {count} cas")
                    
                    problematiques = df_optimise[
                        df_optimise[COL_SALLE_OPTIM] == "Aucune salle adaptée"
                    ]
                    if not problematiques.empty:
                        st.subheader("⚠️ Réservations non attribuées")
                        st.dataframe(
                            problematiques[[COL_DATE, COL_HEURE_DEBUT, COL_HEURE_FIN, 
                                          COL_NB_INSCRITS, COL_CODE_ANALYTIQUE, COL_RAISON_NA]],
                            hide_index=True
                        )
            
            if nb_non_attrib > 0:
                st.warning(t["non_attrib"].format(n=nb_non_attrib))
                
                st.subheader("💡 Recommandations")
                
                causes_principales = df_optimise[
                    df_optimise[COL_SALLE_OPTIM] == "Aucune salle adaptée"
                ][COL_RAISON_NA].value_counts()
                
                if "Capacité insuffisante" in causes_principales:
                    st.info("🔧 **Capacité insuffisante** : Envisagez d'ajouter des salles plus grandes ou de diviser les groupes importants.")
                
                if "Conflit horaire" in causes_principales:
                    st.info("⏰ **Conflits horaires** : Réduisez le buffer ou reprogrammez certains créneaux.")
                
                if not reservations_attribuees.empty:
                    taux_moyen_reel = reservations_attribuees[COL_TAUX_OCCUP].mean()
                    if taux_moyen_reel < 0.6:
                        st.info("📊 **Optimisation possible** : Le taux d'occupation moyen est faible. Considérez la réduction du nombre de salles ou l'augmentation des effectifs.")
            
            st.markdown("---")
            st.subheader("📥 Export des résultats")
            
            try:
                buffer_excel = exporter_excel(df_optimise, seuil_bon, seuil_bas)
                
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                nom_fichier = f"reservations_optimisees_{timestamp}.xlsx"
                
                st.download_button(
                    label=t["telecharger"],
                    data=buffer_excel.getvalue(),
                    file_name=nom_fichier,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary"
                )
                
                st.success("✅ Fichier prêt pour le téléchargement avec formatage conditionnel")
                
            except Exception as e:
                st.error(f"❌ Erreur lors de la génération du fichier Excel : {e}")
        
        else:
            st.error(t["erreur_optim"])
    
    st.markdown("---")
    with st.expander("ℹ️ Informations sur les algorithmes"):
        st.markdown("""
        ### 🤖 Algorithme Glouton Intelligent
        - **Principe** : Traite les réservations par ordre de priorité (durée courte, effectif important)
        - **Avantages** : Rapide, optimal localement, gestion des conflits horaires et doublons
        - **Recommandé pour** : Données avec peu de conflits, optimisation rapide
        
        ### 🧬 Algorithme Génétique
        - **Principe** : Évolution d'une population de solutions par sélection, croisement et mutation
        - **Avantages** : Exploration globale, meilleure optimisation sur données complexes
        - **Recommandé pour** : Données avec beaucoup de contraintes, optimisation fine
        
        ### 📊 Métriques de qualité
        - **Taux optimal** : Occupation entre seuil bas et seuil bon
        - **Sous-utilisation** : Occupation en dessous du seuil bas
        - **Sur-occupation** : Occupation au-dessus du seuil bon (acceptable mais à surveiller)
        
        ### 🔎 Gestion des doublons
        - **Détection** : Identifie les réservations simultanées par CodeAnalytique
        - **Correction** : Conserve la réservation avec le plus grand nombre d'inscrits
        - **Comptage** : Compte les inscrits une seule fois par CodeAnalytique
        """)
    
    st.markdown("---")
    st.markdown(
        "Avec Optimium l'optimisation est à son maximum "
        f"Version 2.1 - {datetime.now().strftime('%Y')}*"
    )
    

if __name__ == "__main__":
    main()